import json
from pymongo import MongoClient
import psycopg2
from psycopg2.extras import execute_values
import hashlib
from pprint import pprint
from datetime import datetime
from openpyxl import Workbook, load_workbook
from itertools import chain
#from xlrd import open_workbook, xldate_as_tuple
import xlrd
from pathlib import Path


MONTHS = [
    "JAN",
    "FEB",
    "MAR",
    "APR",
    "MAY",
    "JUN",
    "JUL",
    "AUG",
    "SEP",
    "OCT",
    "NOV",
    "DEC",
]


def write_shipping_schedule(root_dir):
    #db = MongoClient('mongodb://localhost:27017').development
    conn = psycopg2.connect(
            host="localhost",
            port="5432",
            dbname="postgres",
            user="postgres",
            password="password")
    cur = conn.cursor()

    schedules = root_dir.joinpath('SCHEDULES')
    configs = Path('./configs')
    files = schedules.glob('**/*.xls*')
    #files = [schedules.joinpath('Weekly Shipping Plan.xls')]

    #db.drop_collection('shipping')

    #cur.execute('CREATE DATABASE python_test'
    cur.execute('DROP TABLE IF EXISTS shipping')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS shipping (
          id            int GENERATED ALWAYS AS IDENTITY PRIMARY KEY,
          ship_date     date,
          part          text,
          qty_order     int,
          order_id      text,
          customer      text,
          part_customer text,
          description   text,
          po            text,
          qty_ship      int,
          rem           int,
          price         real,
          vend          text,
          meta          json
        );
    ''');
    conn.commit()

    for filepath in files:
        if not filepath.is_file():
            continue

        p = configs.joinpath(filepath.relative_to(schedules))
        config_filepath = p.parent / (p.name + '.json')
 
        if not (config_filepath.is_file() or config_filepath.is_symlink()):
            continue
 
        parse_shipping_schedule(filepath, config_filepath)

        keys = ["ship_date", "part", "qty_order", "order_id", "customer",
                "part_customer", "description", "po", "qty_ship", "rem",
                "price", "vend", "meta"]

        execute_values(cur, f'INSERT INTO shipping ({", ".join(keys)}) VALUES %s', [[json.dumps(v) if isinstance((v := datum.get(k)), dict) else v for k in keys] for datum in data])
        conn.commit()


def parse_shipping_schedule(filepath: Path, config_filepath: Path):
    with open(config_filepath, 'r') as json_file:
        try:
            config = json.load(json_file)
        except Exception as e:
            raise Exception(f'invalid config file: {config_filepath} ({e})')
    
    wb_config = config['workbook']
    
    default_rows = wb_config.get('rows', [0])
    columns = wb_config['columns']
    
    if filepath.name.endswith('.xls'):
        filehash = hash(filepath)
        wb = xlrd.open_workbook(filepath)
    
        if 'sheets' in wb_config:
            sheets = [(sheet_name, wb.sheet_by_name(sheet_name), sheet_config.get('rows', default_rows))
                    for sheet_name, sheet_config in wb_config['sheets'].items()]
        else:
            # default use only month names
            sheets = [(name, sheet, default_rows)
                    for i in range(wb.nsheets)
                    if any((name := (sheet := wb.sheet_by_index(i)).name).upper().startswith(abrv) for abrv in MONTHS)]
    
            if len(set(s[0] for s in sheets)) != 12:
                raise Exception('missing sheets!')
    
        # iterate sheets
        data = []
        for sheet_name, sheet, rows in sheets:
            print(config_filepath, sheet_name)
            meta = {
                'filename': str(filepath.relative_to(schedules)),
                'sheet':    sheet.name,
                'filehash': filehash,
            }
            # iterate valid sheets
    
            for i in range(0, len(rows), 2):
                # if len(rows) is odd, end_row (which is exclusive) is None
                start_row, end_row = [*rows[i:i+2], None][0:2]
                end_row = end_row or sheet.nrows
    
                print(f'{start_row=} {end_row=}')
                for j in range(start_row, end_row):
                    datum = []
                    for si, column in columns.items():
                        key, col, value_type = (column[k] for k in ('key', 'col', 'value_type'))
                        i = int(si)
                        assert ord(col) - 65 == i, f'invalid col check value {col=} != {i=}'
                        cell = sheet.cell(j, i)
                        _type = sheet.cell_type(j, i)
                        if value_type == 'date':
                            if _type != 3:
                                continue
                            try:
                                tup = xlrd.xldate_as_tuple(cell.value, wb.datemode)
                                dt = datetime(*tup[:3])
                            except:
                                continue
                            datum.append((key, dt))
                        elif value_type == 'int':
                            if _type != 2: # number
                                #raise Exception(f'unknown cell type for int: "{_type}"')
                                continue
                            datum.append((key, int(cell.value)))
                        elif value_type == 'float':
                            if _type != 2: # number
                                #raise Exception(f'unknown cell type for float: "{_type}"')
                                continue
                            datum.append((key, float(cell.value)))
                        elif value_type == 'string':
                            value = cell.value
                            if isinstance(value, float) and value.is_integer():
                                value = int(value)
                            datum.append((key, str(value)))
                        else:
                            raise Exception(f'unknown config value_type "{value_type}"')
                    # if row has more than half the expected values, then keep it
                    if len([d for d in datum if d[1]]) > len(columns) / 2:
                        datum = {k: v for k, v in [*datum, ('meta', meta)]}
                        yield datum
        
    elif filepath.name.endswith('.xlsx'):
        pass
        #f.write(f'{filepath}\n')
        wb = load_workbook(filename=filepath)
        #for i, sheet in enumerate(wb):
        #    f.write(f'sheet "{sheet.title}" {i} rows: {sheet.max_row}\n')
    else:
        continue


def hash(fname):
    h = hashlib.sha1()
    with open(fname, 'rb') as f:
        h.update(f.read())
    return h.hexdigest()


if __name__ == '__main__':
    root_dir = Path.home().joinpath('f')
    if not root_dir.is_dir():
        raise Exception('root_dir is invalid')
    main(root_dir)
